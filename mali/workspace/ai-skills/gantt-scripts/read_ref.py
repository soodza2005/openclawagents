import openpyxl
wb = openpyxl.load_workbook('D:/AI_Agent/openclaw/agents/mali/workspace/Nathalin_Gantt_2025_ref.xlsx', data_only=True)
for name in wb.sheetnames:
    print(f"=== {name} ===")
ws = wb[name]
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10)):
    vals = [str(c.value)[:10] if c.value else '' for c in row]
    print(f"R{i+1}: {vals}")
