import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime, timedelta

wb = openpyxl.Workbook()

# === SHEET 1: DATA ===
ws_data = wb.active
ws_data.title = 'Data'

ws_data['A1'] = 'Week'
ws_data['B1'] = 'Start'
ws_data['C1'] = 'End'

master_start = datetime(2025, 1, 1)
for w in range(52):
    week_s = master_start + timedelta(weeks=w)
    week_e = week_s + timedelta(days=6)
    ws_data[f'A{w+2}'] = f'W{w+1}'
    ws_data[f'B{w+2}'] = week_s
    ws_data[f'C{w+2}'] = week_e

# === SHEET 2: GANTT CHART ===
ws = wb.create_sheet('Gantt Chart')

# Styles
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
label_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
gray_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

# Title
ws.merge_cells('A1:G1')
ws['A1'] = 'GANTT CHART - NATHALIN 2025'
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal='center')

# Headers row 3
headers = ['WBS', 'TASK', 'OWNER', 'START DATE', 'DUE DATE', 'DURATION', '%']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Week timeline
week_col_start = 8

# Row 3: Week header
for w in range(52):
    col = week_col_start + w
    week_s = master_start + timedelta(weeks=w)
    header = f'{week_s.strftime("%d/%m")}'
    cell = ws.cell(row=3, column=col, value=header)
    cell.fill = gray_fill
    cell.font = Font(bold=True, size=7)
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = 5

# Column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 8

# กำหนดช่วงเวลาแต่ละ Project + ชื่อ sub-tasks ตามที่พี่อุ้ยให้มา
projects = {
    '1': {'name': 'Intranet JD (ปรับแก้)', 'start': '01/01/2025', 'end': '10/02/2025', 
          'subs': ['สร้าง Setup เพื่อทำ auto generate Prefix Code', 'สร้าง Setup Main Job', 'ปรับแก้หน้า Menu Main Job manage']},
    '2': {'name': 'Report Financial', 'start': '15/01/2025', 'end': '31/05/2025',
          'subs': ['เก็บข้อมูล', 'Report P&L', 'Report BS', 'Power point Report']},
    '3': {'name': 'Financial Correction', 'start': '01/05/2025', 'end': '30/07/2025',
          'subs': ['เก็บข้อมูล', 'plan project', 'สร้างหน้า รายงานข้อมูลดิบ Fin Dimension', 'สร้างหน้า ตรวจ/ปรับปรุง Fin Dimension App', 'สร้างหน้า สถานะการปิดบัญชี']},
    '4': {'name': 'SitePlan (แผน WFH)', 'start': '15/07/2025', 'end': '30/07/2025',
          'subs': ['หน้าจัดการแผน WFH', 'Dash Board', 'ปรับแก้']},
    '5': {'name': 'Closeaccount', 'start': '15/07/2025', 'end': '31/12/2025',
          'subs': ['เก็บข้อมูล', 'Plan แผนงาน', 'Raw DATA', 'MR Report :Profit & Loss Report', 'MR Report :Balance Sheet Report', 'MR Report :Trial Balance Report', 'MR Report :Standart Report']},
}

# สร้าง tasks ทั้งหมด
all_tasks = []
for proj_num, proj in projects.items():
    proj_start = datetime.strptime(proj['start'], '%d/%m/%Y')
    proj_end = datetime.strptime(proj['end'], '%d/%m/%Y')
    total_days = (proj_end - proj_start).days + 1
    num_tasks = len(proj['subs'])
    days_per_task = total_days // num_tasks
    
    # Project header
    all_tasks.append((proj_num, proj['name'], proj['start'], proj['end']))
    
    # Sub-tasks with names
    task_start = proj_start
    for i, sub_name in enumerate(proj['subs'], 1):
        if i < num_tasks:
            task_end = task_start + timedelta(days=days_per_task - 1)
        else:
            task_end = proj_end
        
        task_name = f'{proj_num}.{i}'
        all_tasks.append((task_name, sub_name, task_start.strftime('%d/%m/%Y'), task_end.strftime('%d/%m/%Y')))
        
        task_start = task_end + timedelta(days=1)

# Colors
color_map = {
    '1': '70AD47',  # Green
    '2': '4472C4',  # Blue
    '3': 'ED7D31',  # Orange
    '4': '7030A0',  # Purple
    '5': '5B9BD5',  # Light Blue
}

row = 4
for wbs, task, start_str, end_str in all_tasks:
    is_main = '.' not in wbs
    
    # WBS
    cell = ws.cell(row=row, column=1, value=wbs)
    if is_main:
        cell.fill = label_fill
    cell.border = thin_border
    
    # Task
    cell = ws.cell(row=row, column=2, value=task)
    if is_main:
        cell.fill = label_fill
    cell.border = thin_border
    
    # Owner
    cell = ws.cell(row=row, column=3, value='')
    cell.border = thin_border
    
    if start_str and end_str:
        start_dt = datetime.strptime(start_str, '%d/%m/%Y')
        end_dt = datetime.strptime(end_str, '%d/%m/%Y')
        
        # START
        cell = ws.cell(row=row, column=4, value=start_dt)
        cell.number_format = 'dd/mm/yyyy'
        cell.border = thin_border
        
        # DUE
        cell = ws.cell(row=row, column=5, value=end_dt)
        cell.number_format = 'dd/mm/yyyy'
        cell.border = thin_border
        
        # DURATION
        cell = ws.cell(row=row, column=6, value=(end_dt - start_dt).days + 1)
        cell.border = thin_border
        
        # %
        cell = ws.cell(row=row, column=7, value=0)
        cell.border = thin_border
        
        # Conditional Formatting
        proj_num = wbs.split('.')[0] if '.' in wbs else wbs
        color = color_map.get(proj_num, '70AD47')
        
        for w in range(52):
            col = week_col_start + w
            col_letter = get_column_letter(col)
            data_row = w + 2
            
            formula = f'AND($D{row}<=Data!$C{data_row},$E{row}>=Data!$B{data_row})'
            
            rule = FormulaRule(
                formula=[formula],
                fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
            )
            
            ws.conditional_formatting.add(f'{col_letter}{row}', rule)
    
    row += 1

wb.save('D:/AI_Agent/openclaw/agents/mali/workspace/Nathalin_Gantt_2025.xlsx')
print('Gantt Chart with full task names created!')
